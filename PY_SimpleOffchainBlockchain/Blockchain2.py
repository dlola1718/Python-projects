import pandas as pd
from Block import Block
import openpyxl
from pathlib import Path

xlsx_file = Path('check_tickers.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 

# print sheet names
# print(wb_obj.sheetnames)


# Read sheet1:

# get_sheet_by_name 
#specify sheet name
sheet_name= "Sheet1"
sheet = wb_obj[sheet_name]
print(sheet)



########uncomment to view datapython
# for row in sheet.iter_rows(): #possible parameters min_row=1, max_col=3, max_row=2 
    # for cell in row:
        # print(cell.value, end=" ")
    # print()


max_col = sheet.max_column
max_row = sheet.max_row
sheet1_transactions = []
sheet1_block = []
for i in range(1, max_row + 1):
    for j in range(1, max_col + 1):
        cell_obj1 = sheet.cell(row = i, column = j) #header then iterate
        # cell_obj2 = sheet.cell(row = 2, column = i)
        # cell_obj3 = sheet.cell(row = 3, column = i)
        genesis_block = Block("Chancellor on the brink...", [ str(cell_obj1.value) ])
        sheet1_transactions.append (genesis_block.block_hash + " ")

        
         
  
#uncomment below to view entire transaction history
# print("Transaction hash: First Sheet")
# print(sheet1_transactions)
 
#data in sheet converted into hash signature
sheet1_block = Block( sheet_name , sheet1_transactions[0::] ) 
print()
print("Block hash: First Block")
print(sheet1_block.block_hash)
print()


    
# blockchain = []

    
#specify sheet name
sheet_name= "Sheet2"
sheet = wb_obj[sheet_name]
# print(sheet)

########uncomment to view data
# for row in sheet.iter_rows(): #possible parameters min_row=1, max_col=3, max_row=2 
#     for cell in row:
#         print(cell.value, end=" ")
#     print()


max_col = sheet.max_column
max_row = sheet.max_row
sheet2_transactions = []
sheet2_block = []
for i in range(1, max_row + 1):
    for j in range(1, max_col + 1):
        cell_obj1 = sheet.cell(row = i, column = j) #header then iterate
        # cell_obj2 = sheet.cell(row = 2, column = i)
        # cell_obj3 = sheet.cell(row = 3, column = i)
        sheet2_genesis_block = Block(sheet1_block.block_hash, [ str(cell_obj1.value) ])
        sheet2_transactions.append (sheet2_genesis_block.block_hash + " ")

#uncomment below to view entire transaction history
# print("Transaction hash: Second Sheet")
# print(sheet1_transactions)
 
#data in sheet converted into hash signature
sheet2_block = Block( sheet_name , sheet2_transactions[0::] ) 
print("Block hash: Second Block")
print(sheet2_block.block_hash)
print()



#specify sheet name
sheet_name= "Sheet3"
sheet = wb_obj[sheet_name]
# print(sheet)

########uncomment to view data
# for row in sheet.iter_rows(): #possible parameters min_row=1, max_col=3, max_row=2 
#     for cell in row:
#         print(cell.value, end=" ")
#     print()


max_col = sheet.max_column
max_row = sheet.max_row
sheet3_transactions = []
sheet3_block = []
for i in range(1, max_row + 1):
    for j in range(1, max_col + 1):
        cell_obj1 = sheet.cell(row = i, column = j) #header then iterate
        # cell_obj2 = sheet.cell(row = 2, column = i)
        # cell_obj3 = sheet.cell(row = 3, column = i)
        sheet3_genesis_block = Block(sheet2_block.block_hash, [ str(cell_obj1.value) ])
        sheet3_transactions.append (sheet3_genesis_block.block_hash + " ")

#uncomment below to view entire transaction history
# print("Transaction hash: Third Sheet")
# print(sheet1_transactions)
 
#data in sheet converted into hash signature
sheet3_block = Block( sheet_name , sheet3_transactions[0::] ) 
print("Block hash: Third Block")
print(sheet3_block.block_hash)
print()




 
