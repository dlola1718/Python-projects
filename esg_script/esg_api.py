"""
ESG Score Searching Tool: This project was Motivated by the REPLY ESG Challenge Event

Imports a list of Stock Tickers from an excel file and exports the esg score results to an excel file 
"""


#ESG Enterprise API Services
#https://www.esgenterprise.com/esg-enterprise-data-api-services/
 
 
import pandas as pd 
from pandas import ExcelWriter   
from pandas import json_normalize 
import requests

import json
import openpyxl
from pathlib import Path
import xlsxwriter     

read_data_path = "check_tickers.xlsx"
output_data_path ="OUTPUT_ESG.xlsx"




######################################################################################################################################################
#load excel file with tickers -------TEMP REDUNDANT
def load_file(self):
    data = pd.read_excel(self, header =0)
    # print(data)
    return data
######################################################################################################################################################


 

######################################################################################################################################################
def save_to_dataframe(the_data):
    df = pd.DataFrame()
    
    df = df.append(the_data)
    df.set_index(df.columns[0], inplace = True)
    print(df)
    writer = pd.ExcelWriter(output_data_path)
    df.to_excel(writer,'ESG Ratings')
    writer.save()
######################################################################################################################################################





######################################################################################################################################################
def check_esg(tickers):
    
    response_list = []
    for i in range(len(tickers)): ##implemented in early stage - temp redundant
        url1 = "https://tf689y3hbj.execute-api.us-east-1.amazonaws.com/prod/authorization/search?q="
        
        url2 = str(tickers)
        url2 = url2.replace("'", "")
        url2 = url2.replace('[','').replace(']','')
        url3 = "&token=d1c32079f707a4a5030aa6a3e514c1cb"
        response = requests.get(url1 + url2 + url3)
        json_data = response.json()
        # df2 = pd.DataFrame()

        # df2 = json_normalize(json_data)


        esg_data = pd.DataFrame(json_data)
        response_list.append(esg_data)
        return response_list
        # return df2
######################################################################################################################################################





######################################################################################################################################################
##READ DATA FROM EXCEL
xlsx_file = Path(read_data_path)
wb_obj = openpyxl.load_workbook(xlsx_file) 

## print sheet names
# print(wb_obj.sheetnames)



## Read sheet1:
# get_sheet_by_name 
##specify sheet name
sheet_name= "Sheet1"
sheet = wb_obj[sheet_name]
# print(sheet)

########uncomment to view data
# for row in sheet.iter_rows(): #possible parameters min_row=1, max_col=3, max_row=2 
#     for cell in row:
#         print(cell.value, end=" ")
#     print()

max_col = sheet.max_column
max_row = 49 ##the demo api supports only 48 queries at a time, so over 1100 stock symbols have been split into 48 entries
sheet1_1_transactions = []
# sheet1_2_transactions = []
# sheet1_3_transactions = []
# sheet1_4_transactions = []
# sheet1_5_transactions = []
# sheet1_6_transactions = []
# sheet1_7_transactions = []
# sheet1_8_transactions = []
# sheet1_9_transactions = []
# sheet1_10_transactions = []
# sheet1_11_transactions = []
# sheet1_12_transactions = []
# sheet1_13_transactions = []
# sheet1_14_transactions = []
# sheet1_15_transactions = []
# sheet1_16_transactions = []
# sheet1_17_transactions = []
# sheet1_18_transactions = []
# sheet1_19_transactions = []
# sheet1_20_transactions = []
# sheet1_21_transactions = []
# sheet1_22_transactions = []
 
for i in range(2, max_row + 1):
    cell_obj1 = sheet.cell(row = i, column = 1) #header then iterate
    # cell_obj2 = sheet.cell(row = i, column = 2)
    # cell_obj3 = sheet.cell(row = i, column = 3)
    # cell_obj4 = sheet.cell(row = i, column = 4)
    # cell_obj5 = sheet.cell(row = i, column = 5)
    # cell_obj6 = sheet.cell(row = i, column = 6)
    # cell_obj7 = sheet.cell(row = i, column = 7)
    # cell_obj8 = sheet.cell(row = i, column = 8)
    # cell_obj9 = sheet.cell(row = i, column = 9)
    # cell_obj10 = sheet.cell(row = i, column = 10)
    # cell_obj11 = sheet.cell(row = i, column = 11)
    # cell_obj12 = sheet.cell(row = i, column = 12)
    # cell_obj13 = sheet.cell(row = i, column = 13)
    # cell_obj14 = sheet.cell(row = i, column = 14)
    # cell_obj15 = sheet.cell(row = i, column = 15)
    # cell_obj16 = sheet.cell(row = i, column = 16)
    # cell_obj17 = sheet.cell(row = i, column = 17)
    # cell_obj18 = sheet.cell(row = i, column = 18)
    # cell_obj19 = sheet.cell(row = i, column = 19)
    # cell_obj20 = sheet.cell(row = i, column = 20)
    # cell_obj21 = sheet.cell(row = i, column = 21)
    # cell_obj22 = sheet.cell(row = i, column = 22)

    sheet1_1_transactions.append (cell_obj1.value)
    # sheet1_2_transactions.append (cell_obj2.value)
    # sheet1_3_transactions.append (cell_obj3.value)
    # sheet1_4_transactions.append (cell_obj4.value)
    # sheet1_5_transactions.append (cell_obj5.value)
    # sheet1_6_transactions.append (cell_obj6.value)
    # sheet1_7_transactions.append (cell_obj7.value)
    # sheet1_8_transactions.append (cell_obj8.value)
    # sheet1_9_transactions.append (cell_obj9.value)
    # sheet1_10_transactions.append (cell_obj10.value)
    # sheet1_11_transactions.append (cell_obj11.value)
    # sheet1_12_transactions.append (cell_obj12.value)
    # sheet1_13_transactions.append (cell_obj13.value)
    # sheet1_14_transactions.append (cell_obj14.value)
    # sheet1_15_transactions.append (cell_obj15.value)
    # sheet1_16_transactions.append (cell_obj16.value)
    # sheet1_17_transactions.append (cell_obj17.value)
    # sheet1_18_transactions.append (cell_obj18.value)
    # sheet1_19_transactions.append (cell_obj19.value)
    # sheet1_20_transactions.append (cell_obj20.value)
    # sheet1_21_transactions.append (cell_obj21.value)
    # sheet1_22_transactions.append (cell_obj22.value)
         
  
##uncomment below to view Ticker symbols 
# print("SYMBOLS")
# print(sheet1_1_transactions[0::])


output_list = []
esg_result = sheet1_1_transactions[0::]
output_list.append(esg_result)
 
save_to_dataframe(check_esg(output_list))
 
# print(check_esg(output_list))

# print(new_df)
# writer = pd.ExcelWriter(output_data_path)
# df.to_excel(writer,'ESG Ratings')
# writer.save()




# save_to_dataframe(check_esg(sheet1_1_transactions[0::]))
# save_to_dataframe(check_esg(sheet1_5_transactions[0::]))

# print(output_list)


# convert to dataframe and save to excel file
# save_to_dataframe(esg_result)