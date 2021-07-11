#ESG Enterprise API Services
#https://www.esgenterprise.com/esg-enterprise-data-api-services/
 
 
from pandas import ExcelWriter    
import requests
import pandas as pd 
import json
import openpyxl
from pathlib import Path
import xlsxwriter     

read_data_path = "check_tickers_copy.xlsx"
output_data_path ="OUTPUT_ESG.xlsx"




######################################################################################################################################################
#load excel file with tickers -------TEMP REDUNDANT
def load_file(self):
    data = pd.read_excel(self, header =0)
    # print(data)
    return data
######################################################################################################################################################


 

######################################################################################################################################################
def save_to_dataframe(the_data):
    df = pd.DataFrame()
    df = df.append(esg_result)
    df.set_index("company_name", inplace = True)
    # print(df)
    writer = pd.ExcelWriter(output_data_path)
    df.to_excel(writer,'ESG Ratings')
    writer.save()
######################################################################################################################################################





######################################################################################################################################################
def check_esg(tickers):
    
    response_list = []
    for i in range(len(tickers)): #implemented in early stage - temp redundant
        url1 = "https://tf689y3hbj.execute-api.us-east-1.amazonaws.com/prod/authorization/search?q="
        
        url2 = str(tickers)
        url2 = url2.replace("'", "")
        url2 = url2.replace('[','').replace(']','')
        url3 = "&token=d1c32079f707a4a5030aa6a3e514c1cb"
        response = requests.get(url1 + url2 + url3)
        json_data = response.json()
        esg_data = pd.DataFrame(json_data)
        response_list.append(esg_data)
        return response_list
######################################################################################################################################################





######################################################################################################################################################
#READ DATA FROM EXCEL
xlsx_file = Path(read_data_path)
wb_obj = openpyxl.load_workbook(xlsx_file) 

# print sheet names
# print(wb_obj.sheetnames)


# Read sheet1:

# get_sheet_by_name 
#specify sheet name
sheet_name= "Sheet1"
sheet = wb_obj[sheet_name]
# print(sheet)

########uncomment to view data
# for row in sheet.iter_rows(): #possible parameters min_row=1, max_col=3, max_row=2 
#     for cell in row:
#         print(cell.value, end=" ")
#     print()

max_col = sheet.max_column
max_row = 49 #the demo api supports only 48 queries at a time
sheet1_transactions = []
sheet1_block = []
cell_obj=[]
for i in range(2, max_row + 1):
    cell_obj[1] = sheet.cell(row = i, column = 1) #header then iterate
    # cell_obj2 = sheet.cell(row = i, column = 2)
    # cell_obj3 = sheet.cell(row = i, column = 3)
    # cell_obj4 = sheet.cell(row = i, column = 4)
    # cell_obj5 = sheet.cell(row = i, column = 5)
    # cell_obj6 = sheet.cell(row = i, column = 6)
    # cell_obj7 = sheet.cell(row = i, column = 7)
    # cell_obj8 = sheet.cell(row = i, column = 8)
    # cell_obj9 = sheet.cell(row = i, column = 9)
    # cell_obj10 = sheet.cell(row = i, column = 10)
    # cell_obj11 = sheet.cell(row = i, column = 11)
    # cell_obj12 = sheet.cell(row = i, column = 12)
    # cell_obj13 = sheet.cell(row = i, column = 13)
    # cell_obj14 = sheet.cell(row = i, column = 14)
    # cell_obj15 = sheet.cell(row = i, column = 15)
    # cell_obj16 = sheet.cell(row = i, column = 16)
    # cell_obj17 = sheet.cell(row = i, column = 17)
    # cell_obj18 = sheet.cell(row = i, column = 18)
    # cell_obj19 = sheet.cell(row = i, column = 19)
    # cell_obj20 = sheet.cell(row = i, column = 20)
    # cell_obj21 = sheet.cell(row = i, column = 21)
    # cell_obj22 = sheet.cell(row = i, column = 22)

    print(cell_obj[1])
        # sheet1_transactions.append (test_cell.value)
    j += 3

        
         
  
#uncomment below to view Ticker symbols 
# print("SYMBOLS")
print(sheet1_transactions[0::])


output_list = []
# esg_result = check_esg(sheet1_transactions[0::])
# output_list.append(esg_result)
# print(output_list)



# convert to dataframe and save to excel file
# save_to_dataframe(esg_result)